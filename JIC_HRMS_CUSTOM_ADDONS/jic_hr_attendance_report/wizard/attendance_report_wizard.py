# -*- coding: utf-8 -*-
import base64
from io import BytesIO

try:
    import openpyxl as xl
except ImportError:
    xl = None

from odoo import api, fields, models, _, SUPERUSER_ID
from odoo.tools import date_utils
from datetime import datetime, date, timedelta, time
from collections import OrderedDict, defaultdict
import calendar
import pytz
from odoo.addons.resource.models.resource import float_to_time
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, NamedStyle
from odoo.addons.jic_base.models import openpyxl_formats as openpyxl_formats


class HrAttendanceReportWizard(models.TransientModel):
    _name = 'hr.attendance.report.wizard'
    _description = 'Attendance Report Wizard'

    def _get_date_from(self):
        date = datetime.today()
        return datetime(date.year, date.month, 1).strftime("%Y-%m-%d")

    def _get_date_to(self):
        date = datetime.today()
        return datetime(date.year, date.month, calendar.mdays[date.month]).strftime("%Y-%m-%d")

    report_file = fields.Binary(string="Report File")

    date_from = fields.Date(string="Date From", required=True, default=_get_date_from)
    date_to = fields.Date(string="Date to", required=True, default=_get_date_to)
    department_id = fields.Many2one("hr.department", string="Department")
    employee_id = fields.Many2one("hr.employee", string="Employee")
    debug = fields.Boolean(string="Debug?")

    def date_range_list(self, start_date, end_date):
        # Return list of datetime.date objects between start_date and end_date (inclusive).
        date_list = []
        curr_date = start_date
        while curr_date <= end_date:
            date_list.append(curr_date)
            curr_date += timedelta(days=1)
        return date_list

    def to_user_tz(self, user_tz, datetime_to_convert):
        """
        This to convert python datetime object to datetime string with timezone
        :param user_tz: timezone
        :param datetime_to_convert: python datetime object
        :return: date string
        """
        dt_obj = datetime_to_convert.astimezone(pytz.timezone(user_tz))
        return fields.Datetime.to_string(dt_obj)

    def to_user_tz_datetime(self, user_tz, datetime_to_convert):
        """
        This to convert python datetime object to datetime with timezone
        :param user_tz: timezone
        :param datetime_to_convert: python datetime object
        :return: datetime object
        """
        dt_obj = datetime_to_convert.astimezone(pytz.timezone(user_tz))
        return fields.Datetime.to_string(dt_obj)

    def find_checkin_checkout_as_per_calendar(self, employee_id, day, half_day, half_day_type):
        start_time = 0
        end_time = 0
        start_time_datetime = day
        end_time_datetime = day
        my_time = datetime.min.time()
        for week_days in employee_id.resource_calendar_id.attendance_ids:
            if int(week_days.dayofweek) == int(day.weekday()):
                if half_day:
                    if half_day_type == 'am':
                        if week_days.day_period == 'morning':
                            start_time = week_days.hour_from
                            end_time = week_days.hour_to
                            continue
                    if half_day_type == 'pm':
                        if week_days.day_period == 'afternoon':
                            start_time = week_days.hour_from
                            end_time = week_days.hour_to
                            continue
                else:
                    if week_days.day_period == 'morning':
                        start_time = week_days.hour_from
                    if week_days.day_period == 'afternoon':
                        end_time = week_days.hour_to
                    continue
        if start_time:
            start_time_datetime = datetime.combine(day, my_time) + timedelta(
                hours=int(str(start_time).split('.')[0]),
                minutes=int(str(start_time).split('.')[1])
            )
        if end_time:
            end_time_datetime = datetime.combine(day, my_time) + timedelta(
                hours=int(str(end_time).split('.')[0]),
                minutes=int(str(end_time).split('.')[1])
            )
        legal_hours = end_time - start_time
        return start_time_datetime, end_time_datetime, legal_hours

    def find_timings(self, employee_id, date_to_check):

        # Start here------------------------------>>>>

        employee_id_obj = self.env["hr.employee"].browse(employee_id)
        department_id = employee_id_obj.department_id
        user_tz = employee_id_obj.tz or self.env.context.get('tz')
        total_worked_hours = 0
        total_hours = 0
        approved_timesheet = 0
        first_check_in_tz = False
        last_check_out_tz = False
        start_time = 0
        end_time = 0
        grace_period = 0
        max_allowed_exceptions = 0

        if department_id:
            grace_period = department_id.grace_period_in_attendance
            max_allowed_exceptions = department_id.max_allowed_exceptions_in_month

        date_to_check_start = fields.Datetime.from_string(date_to_check.strftime("%Y-%m-%d 00:00:00"))
        date_to_check_stop = date_to_check_start + timedelta(days=1, seconds=-1)

        #date_to_check_start_user_tz = self.to_user_tz(user_tz, date_to_check_start)
        #date_to_check_stop_user_tz = self.to_user_tz(user_tz, date_to_check_stop)

        # Find attendance
        attendance_a_day_ids = self.env["hr.attendance"].search(
            [
                ("check_in",">=", date_to_check_start),
                ("check_out","<=", date_to_check_stop),
                ("employee_id", "=", employee_id)
            ]
        )
        # Remove all the attendance without check out
        attendance_a_day_ids -= attendance_a_day_ids.filtered(lambda a: not a.check_out)
        if attendance_a_day_ids:
            first_check_in = min(attendance_a_day_ids.mapped("check_in") if attendance_a_day_ids else [])
            first_check_in_tz = self.to_user_tz(user_tz, fields.Datetime.from_string(first_check_in))
            last_check_out = max(attendance_a_day_ids.mapped("check_out") if attendance_a_day_ids else [])
            last_check_out_tz = self.to_user_tz(user_tz, fields.Datetime.from_string(last_check_out))
            total_hours = (last_check_out - first_check_in).total_seconds() / 3600
            total_worked_hours = sum(attendance_a_day_ids.mapped("worked_hours"))

        # Find timesheet
        date_timesheet = fields.Date.from_string(date_to_check)
        timesheet_a_day_ids = self.env["account.analytic.line"].search(
            [
                ("date", "=", date_timesheet),
                ("employee_id", "=", employee_id)
            ]
        )
        if timesheet_a_day_ids:
            approved_timesheet = sum(timesheet_a_day_ids.mapped("unit_amount"))

        resource_id = employee_id_obj.resource_id

        day_from = datetime.combine(date_to_check_start, time.min).replace(tzinfo=pytz.UTC)
        day_to = datetime.combine(date_to_check_stop, time.max).replace(tzinfo=pytz.UTC)

        # Find Leave hours
        domain = [('time_type', '=', 'leave')]
        leave_intervals = employee_id_obj.resource_calendar_id._leave_intervals_batch(
            day_from, day_to, resources=[resource_id], domain=domain)[resource_id.id]

        leave_hours = 0
        half_day = False
        half_day_type = 'am'
        for start, stop, leave_id in leave_intervals:
            leave_hours += (stop - start).total_seconds() / 3600
            if leave_id.holiday_id:
                if leave_id.holiday_id.request_unit_half:
                    half_day = True
                    if leave_id.holiday_id.request_date_from_period == 'pm':
                        half_day_type = 'pm'

        start_time, end_time, legal_hours = self.find_checkin_checkout_as_per_calendar(employee_id_obj, date_to_check, half_day, half_day_type)

        # Find legal hours - as per the contract
        global_leaves = employee_id_obj.resource_calendar_id._attendance_intervals_batch(day_from, day_to, resources=[resource_id])[resource_id.id] \
                    - employee_id_obj.resource_calendar_id._leave_intervals_batch(day_from, day_to, None)[False]  # Substract Global Leaves
        contract_hours = sum((stop - start).total_seconds() / 3600 for start, stop, dummy in global_leaves)
        legal = employee_id_obj.list_work_time_per_day(day_from, day_to, None)
        legal_hours = legal[0][1] if legal else 0

        # Find late in and early out
        late_hr = 0
        late_min = 0
        early_hr = 0
        early_min = 0
        first_chk_in = fields.Datetime.from_string(first_check_in_tz)
        last_chk_out = fields.Datetime.from_string(last_check_out_tz)
        if start_time and first_chk_in:
            if start_time < first_chk_in:
                late_hours = first_chk_in - start_time
                late_hr = int(str(late_hours).split(":")[0])
                late_min = int(str(late_hours).split(":")[1])
        if end_time and last_chk_out:
            if last_chk_out < end_time:
                early_hours = end_time - last_chk_out
                early_hr = int(str(early_hours).split(":")[0])
                early_min = int(str(early_hours).split(":")[1])

        return {
            "calendar_hours": contract_hours or 0,                  # Global contract hours
            "leave_hours": leave_hours or 0,                        # in case of leave taken
            "total_hours": total_hours or 0,                        # Based on first in and last out
            "total_worked_hours": total_worked_hours or 0,          # Real attendance
            "break_time": round(total_hours - total_worked_hours) or 0,  # Break only
            "first_in": first_check_in_tz or
                        date_to_check.strftime("%Y-%m-%d 00:00:00"),
            "last_out": last_check_out_tz or
                        date_to_check.strftime("%Y-%m-%d 00:00:00"),
            "approved_timesheet": approved_timesheet or 0,
            "half_day": half_day,
            "half_day_type": half_day_type,
            "real_start_time": start_time,
            "real_end_time": end_time,
            "late_hr": late_hr,
            "late_min": late_min,
            "early_hr": early_hr,
            "early_min": early_min,
            "grace_period": grace_period,
            "max_allowed_exceptions": max_allowed_exceptions,
            "legal_hours": legal_hours
        }

    def get_data_to_plot(self):
        date_from = self.date_from
        date_to = self.date_to

        date_list = self.date_range_list(date_from, date_to)
        attendance_lines = {}

        # Prepare domain
        domain = [("user_id", "!=", 2), ("opt_for_attendance_mail", "=", True)]
        if self.department_id:
            domain.append(("department_id", "=", self.department_id.id))
        if self.employee_id:
            domain.append(("id", "=", self.employee_id.id))

        # Get employees\
        employees_list = self.env["hr.employee"].search_read(
            domain,
            {"id", "name", "emp_code", "job_title", "mobile_phone", "work_email", "user_id"}
        )

        dict_of_dates = dict.fromkeys(date_list, 1)

        for date_attendance in dict_of_dates:
            list_temp = []
            for employee in employees_list:

                time_data = self.find_timings(employee.get("id"), date_attendance)

                if time_data.get("calendar_hours") and time_data.get("calendar_hours") <= time_data.get("leave_hours"):
                    attendance_status = "LEAVE-F"
                elif time_data.get("leave_hours") and time_data.get("calendar_hours") > time_data.get("leave_hours"):
                    attendance_status = "LEAVE-H"
                else:
                    attendance_status = "PRESENT"

                holiday = False
                if not time_data.get("calendar_hours"):
                    holiday = True
                    attendance_status = "OFF"

                # Overtime hours
                overtime_hours = abs(time_data.get("total_worked_hours") - time_data.get("legal_hours")) \
                                if (time_data.get("total_worked_hours") > time_data.get("legal_hours")) else 0

                # Shortage hours
                shortage = (
                    time_data.get("calendar_hours") - time_data.get("leave_hours")
                           ) - time_data.get("total_worked_hours")
                shortage = shortage if shortage > 0 else 0

                attendance_line = {
                    "date": date_attendance,
                    "employee_id": employee.get("id"),
                    "employee_name": employee.get("name"),
                    "in_time": time_data.get("first_in"),
                    "real_start_time": time_data.get("real_start_time"),
                    "in_time_str": self.datetime_to_hour_string(time_data.get("first_in")),
                    "out_time": time_data.get("last_out"),
                    "real_end_time": time_data.get("real_end_time"),
                    "out_time_str": self.datetime_to_hour_string(time_data.get("last_out")),
                    "total_time": time_data.get("total_hours"),         # Based on first in and last out
                    "total_time_str": self.float_to_time_excel(time_data.get("total_hours")),
                    "break_time": time_data.get("break_time"),          # Break only
                    "break_time_str": self.float_to_time_excel(time_data.get("break_time")),
                    "net_time": time_data.get("total_worked_hours"),    # Net worked time
                    "net_time_str": self.float_to_time_excel(time_data.get("total_worked_hours")),
                    "calendar_hours": time_data.get("calendar_hours"),  # Calendar hour as per contract .ie full time (8hrs)
                    "calendar_hours_str": self.float_to_time_excel(time_data.get("calendar_hours")),
                    "legal_hours": time_data.get("legal_hours"),  # Legal hours to work for a day (8hrs - leave hours)
                    "legal_hours_str": self.float_to_time_excel(time_data.get("legal_hours")),
                    "overtime_hours": overtime_hours,
                    "overtime_hours_str": self.float_to_time_excel(time_data.get("overtime_hours")),
                    "leave_hours": time_data.get("leave_hours"),        # Leave hours
                    "leave_hours_str": self.float_to_time_excel(time_data.get("leave_hours")),
                    "shortage_hours": shortage,
                    "shortage_hours_str": self.float_to_time_excel(shortage),
                    "attendance_status": attendance_status,
                    "approved_timesheet_hours": time_data.get("approved_timesheet"),
                    "approved_timesheet": self.float_to_time_excel(time_data.get("approved_timesheet")) or 0,
                    "holiday": holiday,
                    "late_hr": time_data.get("late_hr"),
                    "late_min": time_data.get("late_min"),
                    "early_hr": time_data.get("early_hr"),
                    "early_min": time_data.get("early_min"),
                    "grace_period": time_data.get("grace_period"),
                    "max_allowed_exceptions": time_data.get("max_allowed_exceptions"),
                    "half_day": time_data.get("half_day"),
                    "half_day_type": time_data.get("half_day_type"),
                }
                list_temp.append(attendance_line)
            attendance_lines.update({date_attendance: list_temp})
        return date_list, attendance_lines, employees_list

    def float_to_time_excel(self, float_value):
        if float_value and float_value <= 23:
            return float_to_time(round(abs(float(float_value)), 2)).strftime("%H:%M") or "00:00"
        else:
            "00:00"

    def float_to_time_excel_unlimited(self, float_value):
        if float_value:
            #return float_to_time(round(abs(float(float_value)), 2)).strftime("%H:%M") or "00:00"
            minutes = float_value * 60
            hours, minutes = divmod(minutes, 60)
            return "%02d:%02d" % (hours, minutes)
        else:
            "00:00"

    def datetime_to_hour_string(self, datetime_string):
        if datetime_string:
            return fields.Datetime.from_string(datetime_string).strftime("%H:%M")
        return "00:00"

    def datetime_to_hour_float(self, datetime_string):
        if datetime_string:
            return fields.Datetime.from_string(datetime_string).strftime("%H.%M")
        return "00:00"

    def plot_data(self, date_list, attendance_lines, employees_list,
                  data_start_row, data_start_column, sheet):

        # Fill date row first
        for date_val in date_list:
            sheet.cell(row=data_start_row, column=data_start_column, value=date_val)
            data_start_column += 1

        data_start_row += 1
        data_start_column = 5

        # Fill date row second
        for date_val in date_list:
            sheet.cell(row=data_start_row, column=data_start_column, value=date_val.strftime("%A"))
            data_start_column += 1

        data_start_row += 1
        data_start_column = 1

        # Fill attendance data

        for employee in employees_list:
            sheet.cell(row=data_start_row, column=data_start_column, value=employee["name"])
            data_start_column += 1
            sheet.cell(row=data_start_row, column=data_start_column, value=employee["job_title"])
            data_start_column += 1
            sheet.cell(row=data_start_row, column=data_start_column, value=employee["emp_code"])
            data_start_column += 1
            sheet.cell(row=data_start_row, column=data_start_column, value="ATTND")
            data_start_column += 1

            data_start_row += 1
            data_start_column = 1

            sheet.cell(row=data_start_row, column=data_start_column, value="")
            data_start_column += 1
            sheet.cell(row=data_start_row, column=data_start_column, value=employee["mobile_phone"])
            data_start_column += 1
            sheet.cell(row=data_start_row, column=data_start_column, value="")
            data_start_column += 1
            sheet.cell(row=data_start_row, column=data_start_column, value="CHECK IN")
            data_start_column += 1

            data_start_row += 1
            data_start_column = 1

            sheet.cell(row=data_start_row, column=data_start_column, value="")
            data_start_column += 1
            sheet.cell(row=data_start_row, column=data_start_column, value=employee["work_email"])
            data_start_column += 1
            sheet.cell(row=data_start_row, column=data_start_column, value="")
            data_start_column += 1
            sheet.cell(row=data_start_row, column=data_start_column, value="CHECK OUT")
            data_start_column += 1

            data_start_row += 1
            data_start_column = 1

            sheet.cell(row=data_start_row, column=data_start_column, value="")
            data_start_column += 1
            sheet.cell(row=data_start_row, column=data_start_column, value="")
            data_start_column += 1
            sheet.cell(row=data_start_row, column=data_start_column, value="")
            data_start_column += 1
            sheet.cell(row=data_start_row, column=data_start_column, value="BREAK")
            data_start_column += 1

            data_start_row += 1
            data_start_column = 1

            sheet.cell(row=data_start_row, column=data_start_column, value="")
            data_start_column += 1
            sheet.cell(row=data_start_row, column=data_start_column, value="")
            data_start_column += 1
            sheet.cell(row=data_start_row, column=data_start_column, value="")
            data_start_column += 1
            sheet.cell(row=data_start_row, column=data_start_column, value="SHORT")
            data_start_column += 1

            data_start_row += 1
            data_start_column = 1

            sheet.cell(row=data_start_row, column=data_start_column, value="")
            data_start_column += 1
            sheet.cell(row=data_start_row, column=data_start_column, value="")
            data_start_column += 1
            sheet.cell(row=data_start_row, column=data_start_column, value="")
            data_start_column += 1
            sheet.cell(row=data_start_row, column=data_start_column, value="LATE IN")
            data_start_column += 1

            data_start_row += 1
            data_start_column = 1

            sheet.cell(row=data_start_row, column=data_start_column, value="")
            data_start_column += 1
            sheet.cell(row=data_start_row, column=data_start_column, value="")
            data_start_column += 1
            sheet.cell(row=data_start_row, column=data_start_column, value="")
            data_start_column += 1
            sheet.cell(row=data_start_row, column=data_start_column, value="EARLY OUT")
            data_start_column += 1

            data_start_row += 1
            data_start_column = 1

            sheet.cell(row=data_start_row, column=data_start_column, value="")
            data_start_column += 1
            sheet.cell(row=data_start_row, column=data_start_column, value="")
            data_start_column += 1
            sheet.cell(row=data_start_row, column=data_start_column, value="")
            data_start_column += 1
            sheet.cell(row=data_start_row, column=data_start_column, value="TIMESHEET")
            data_start_column += 1

            data_start_row += 1
            data_start_column = 1

            sheet.cell(row=data_start_row, column=data_start_column, value="")
            data_start_column += 1
            sheet.cell(row=data_start_row, column=data_start_column, value="")
            data_start_column += 1
            sheet.cell(row=data_start_row, column=data_start_column, value="")
            data_start_column += 1
            net_hours = sheet.cell(row=data_start_row, column=data_start_column, value="NH")
            data_start_column += 1
            net_hours.style = openpyxl_formats.format_bold_no_color_special

            data_start_row -= 8
            data_start_column = 5

            for date_val in date_list:
                for attendance_line in attendance_lines.get(date_val):
                    if attendance_line.get("employee_id") == employee["id"]:
                        attn_status = sheet.cell(row=data_start_row, column=data_start_column, value=attendance_line.get("attendance_status"))
                        attn_status.style = openpyxl_formats.format_unbold_data_default
                        if attendance_line.get("attendance_status") != 'PRESENT' or \
                                attendance_line.get("shortage_hours") >= 4 or \
                                attendance_line.get("shortage_hours"):
                            attn_status.style = openpyxl_formats.format_leave
                        data_start_column += 1
                        if attendance_line.get("holiday"):
                            attn_status.style = openpyxl_formats.format_unbold_with_color

            data_start_column = 5
            data_start_row += 1

            for date_val in date_list:
                for attendance_line in attendance_lines.get(date_val):
                    if attendance_line.get("employee_id") == employee["id"]:
                        in_time = sheet.cell(row=data_start_row, column=data_start_column, value=self.datetime_to_hour_string(attendance_line.get("in_time")))
                        in_time.style = openpyxl_formats.format_unbold_data_default
                        data_start_column += 1
                        if attendance_line.get("holiday"):
                            in_time.style = openpyxl_formats.format_unbold_with_color

            data_start_column = 5
            data_start_row += 1

            for date_val in date_list:
                for attendance_line in attendance_lines.get(date_val):
                    if attendance_line.get("employee_id") == employee["id"]:
                        out_time = sheet.cell(row=data_start_row, column=data_start_column, value=self.datetime_to_hour_string(attendance_line.get("out_time")))
                        out_time.style = openpyxl_formats.format_unbold_data_default
                        data_start_column += 1
                        if attendance_line.get("holiday"):
                            out_time.style = openpyxl_formats.format_unbold_with_color
            data_start_column = 5
            data_start_row += 1

            for date_val in date_list:
                for attendance_line in attendance_lines.get(date_val):
                    if attendance_line.get("employee_id") == employee["id"]:
                        break_time = sheet.cell(row=data_start_row, column=data_start_column, value=self.float_to_time_excel(attendance_line.get("break_time", 0)))
                        break_time.style = openpyxl_formats.format_unbold_data_default
                        data_start_column += 1
                        if attendance_line.get("holiday"):
                            break_time.style = openpyxl_formats.format_unbold_with_color

            data_start_column = 5
            data_start_row += 1

            total_shortage_hrs_count = 0
            for date_val in date_list:
                for attendance_line in attendance_lines.get(date_val):
                    if attendance_line.get("employee_id") == employee["id"]:
                        shortage_hours = sheet.cell(row=data_start_row, column=data_start_column, value=self.float_to_time_excel(attendance_line.get("shortage_hours",0)))
                        if attendance_line.get("shortage_hours",0):
                            shortage_hours.style = openpyxl_formats.format_unbold_with_color_danger
                        else:
                            shortage_hours.style = openpyxl_formats.format_unbold_data_default
                        total_shortage_hrs_count += attendance_line.get("shortage_hours",0)
                        data_start_column += 1
                        if attendance_line.get("holiday"):
                            shortage_hours.style = openpyxl_formats.format_unbold_with_color

            total_shortage = sheet.cell(row=data_start_row, column=data_start_column,
                                    value=self.float_to_time_excel_unlimited(total_shortage_hrs_count))
            total_shortage.style = openpyxl_formats.format_final_counts

            data_start_column = 5
            data_start_row += 1

            total_late_hrs_count = 0
            for date_val in date_list:
                for attendance_line in attendance_lines.get(date_val):
                    if attendance_line.get("employee_id") == employee["id"]:
                        if attendance_line.get("late_hr", 0) or attendance_line.get(
                                "late_min", 0):
                            late_hours = sheet.cell(row=data_start_row, column=data_start_column,
                                value=str(attendance_line.get("late_hr", 0)) + ":" + str(attendance_line.get("late_min", 0)))
                            late_hours.style = openpyxl_formats.format_unbold_with_color_danger
                            total_late_hrs_count += 1
                        else:
                            late_hours = sheet.cell(row=data_start_row, column=data_start_column,
                                                    value="")
                            late_hours.style = openpyxl_formats.format_unbold_data_default
                        data_start_column += 1
                        if attendance_line.get("holiday"):
                            late_hours.style = openpyxl_formats.format_unbold_with_color

            total_late = sheet.cell(row=data_start_row, column=data_start_column,
                                     value=total_late_hrs_count)
            total_late.style = openpyxl_formats.format_final_counts

            data_start_column = 5
            data_start_row += 1

            total_early_hrs_count = 0
            for date_val in date_list:
                for attendance_line in attendance_lines.get(date_val):
                    if attendance_line.get("employee_id") == employee["id"]:
                        if attendance_line.get("early_hr", 0) or attendance_line.get(
                                                        "early_min", 0):
                            early_hours = sheet.cell(row=data_start_row, column=data_start_column,
                                value=str(attendance_line.get("early_hr", 0)) + ":" + str(attendance_line.get(
                                                            "early_min", 0)))
                            early_hours.style = openpyxl_formats.format_unbold_with_color_danger
                            total_early_hrs_count += 1
                        else:
                            early_hours = sheet.cell(row=data_start_row, column=data_start_column,
                                                     value="")
                            early_hours.style = openpyxl_formats.format_unbold_data_default
                        data_start_column += 1
                        if attendance_line.get("holiday"):
                            early_hours.style = openpyxl_formats.format_unbold_with_color
            total_early = sheet.cell(row=data_start_row, column=data_start_column,
                                         value=total_early_hrs_count)
            total_early.style = openpyxl_formats.format_final_counts

            data_start_column = 5
            data_start_row += 1

            total_approved_timesheet = 0
            for date_val in date_list:
                for attendance_line in attendance_lines.get(date_val):
                    if attendance_line.get("employee_id") == employee["id"]:
                        timesheet_hours = sheet.cell(row=data_start_row, column=data_start_column,
                                                    value=self.float_to_time_excel(
                                                        attendance_line.get("approved_timesheet_hours", 0)))
                        timesheet_hours.style = openpyxl_formats.format_unbold_data_default
                        total_approved_timesheet += attendance_line.get("approved_timesheet_hours", 0)
                        data_start_column += 1
                        if attendance_line.get("holiday"):
                            timesheet_hours.style = openpyxl_formats.format_unbold_with_color

            total_timesheet = sheet.cell(row=data_start_row, column=data_start_column,
                       value=self.float_to_time_excel_unlimited(total_approved_timesheet))
            total_timesheet.style = openpyxl_formats.format_final_counts

            data_start_column = 5
            data_start_row += 1

            total_net_hours = 0
            for date_val in date_list:
                for attendance_line in attendance_lines.get(date_val):
                    if attendance_line.get("employee_id") == employee["id"]:
                        net_time = sheet.cell(row=data_start_row, column=data_start_column, value=self.float_to_time_excel(attendance_line.get("net_time",0)))
                        data_start_column += 1
                        net_time.style = openpyxl_formats.format_bold_no_color
                        total_net_hours += attendance_line.get("net_time",0)
                        if attendance_line.get("holiday"):
                            net_time.style = openpyxl_formats.format_unbold_with_color

            total_net = sheet.cell(row=data_start_row, column=data_start_column,
                                         value=self.float_to_time_excel_unlimited(total_net_hours))
            total_net.style = openpyxl_formats.format_final_counts

            data_start_column = 1
            data_start_row += 1

            # Blank row
            max_col = len(date_list) + 4
            for rows in sheet.iter_rows(min_row=data_start_row, max_row=data_start_row, min_col=1, max_col=max_col):
                for cell in rows:
                    cell.style = openpyxl_formats.format_vaccant_row
            data_start_row += 1

    def plot_all_data_to_debug(self, sheet, attendance_lines):
        sheet2_row = 1
        sheet2_col = 1

        sheet.cell(row=sheet2_row, column=sheet2_col + 1, value="Date").style = openpyxl_formats.format_table_head
        sheet.cell(row=sheet2_row, column=sheet2_col + 2, value="Employee").style = openpyxl_formats.format_table_head
        sheet.cell(row=sheet2_row, column=sheet2_col + 3, value="In Time").style = openpyxl_formats.format_table_head
        sheet.cell(row=sheet2_row, column=sheet2_col + 4, value="Out Time").style = openpyxl_formats.format_table_head
        sheet.cell(row=sheet2_row, column=sheet2_col + 5, value="Break Time").style = openpyxl_formats.format_table_head
        sheet.cell(row=sheet2_row, column=sheet2_col + 6, value="Net Time").style = openpyxl_formats.format_table_head
        sheet.cell(row=sheet2_row, column=sheet2_col + 7, value="Contract Hours").style = openpyxl_formats.format_table_head
        sheet.cell(row=sheet2_row, column=sheet2_col + 8, value="Leave Hours").style = openpyxl_formats.format_table_head
        sheet.cell(row=sheet2_row, column=sheet2_col + 9, value="Shortage Hours").style = openpyxl_formats.format_table_head
        sheet.cell(row=sheet2_row, column=sheet2_col + 10, value="Status").style = openpyxl_formats.format_table_head
        sheet.cell(row=sheet2_row, column=sheet2_col + 11, value="Is Holiday").style = openpyxl_formats.format_table_head

        sheet2_row = 1
        for att_line in attendance_lines.keys():
            sheet.cell(row=sheet2_row, column=sheet2_col, value=att_line)
            sheet2_row += 1
            for r in attendance_lines[att_line]:
                sheet.cell(row=sheet2_row, column=sheet2_col + 1, value=r.get("date"))
                sheet.cell(row=sheet2_row, column=sheet2_col + 2, value=r.get("employee_name"))
                sheet.cell(row=sheet2_row, column=sheet2_col + 3, value=self.datetime_to_hour_string(r.get("in_time")))
                sheet.cell(row=sheet2_row, column=sheet2_col + 4, value=self.datetime_to_hour_string(r.get("out_time")))
                sheet.cell(row=sheet2_row, column=sheet2_col + 5, value=self.float_to_time_excel(r.get("break_time")))
                sheet.cell(row=sheet2_row, column=sheet2_col + 6, value=self.float_to_time_excel(r.get("net_time")))
                sheet.cell(row=sheet2_row, column=sheet2_col + 7, value=self.float_to_time_excel(r.get("calendar_hours")))
                sheet.cell(row=sheet2_row, column=sheet2_col + 8, value=self.float_to_time_excel(r.get("leave_hours")))
                sheet.cell(row=sheet2_row, column=sheet2_col + 9, value=self.float_to_time_excel(r.get("shortage_hours")))
                sheet.cell(row=sheet2_row, column=sheet2_col + 10, value=r.get("attendance_status"))
                sheet.cell(row=sheet2_row, column=sheet2_col + 11, value=r.get("holiday"))


                sheet2_row += 1

    def prepare_column_width(self, sheet):

        # Set row height
        sheet.row_dimensions[1].height = 30
        sheet.row_dimensions[2].height = 20

        # if we need to add alignment to single cell
        # sheet.cell(3,1).alignment = Alignment(horizontal='center', vertical='center')

        # Set alignment to row 3
        for cell in sheet["3:3"]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(b=True)

        # Set alignment to row 4
        for cell in sheet["4:4"]:
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Set column width
        sheet.column_dimensions[get_column_letter(1)].width = 30
        sheet.column_dimensions[get_column_letter(2)].width = 30
        sheet.column_dimensions[get_column_letter(3)].width = 30

        for i in range(4, 60):
            sheet.column_dimensions[get_column_letter(i)].width = 15

        for i in range(1, 60):
            sheet.column_dimensions[get_column_letter(i)].alignment = Alignment(horizontal='center', vertical='center')

    def action_download(self):

        data_start_row = 3
        data_start_column = 5

        attachment = self.env.ref(
            "jic_hr_attendance_report.jic_attendance_report_format", raise_if_not_found=True
        )

        wb = xl.load_workbook(BytesIO(base64.decodebytes(attachment.datas)))
        ws = wb.active
        ws.sheet_view.showGridLines = False
        ws.sheet_view.windowProtection = True

        ws.title = "Attendance Summary"
        ws.sheet_view.zoomScale = 80
        ws.freeze_panes = 'E5'

        # Prepare width of columns
        self.prepare_column_width(ws)
        date_list, attendance_lines, employees_list = self.get_data_to_plot()
        self.plot_data(date_list, attendance_lines, employees_list, data_start_row, data_start_column, ws)

        if self.debug:

            # New sheet for debug
            ws2 = wb.create_sheet("Debug")

            # Set column width
            ws2.column_dimensions[get_column_letter(1)].width = 15
            ws2.column_dimensions[get_column_letter(2)].width = 15
            ws2.column_dimensions[get_column_letter(3)].width = 30
            ws2.column_dimensions[get_column_letter(4)].width = 15
            ws2.column_dimensions[get_column_letter(5)].width = 15
            ws2.column_dimensions[get_column_letter(6)].width = 15
            ws2.column_dimensions[get_column_letter(7)].width = 18
            ws2.column_dimensions[get_column_letter(8)].width = 18
            ws2.column_dimensions[get_column_letter(9)].width = 18
            ws2.column_dimensions[get_column_letter(10)].width = 18
            ws2.column_dimensions[get_column_letter(11)].width = 15
            ws2.column_dimensions[get_column_letter(12)].width = 15

            self.plot_all_data_to_debug(ws2, attendance_lines)

        # Finalize and download the file

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        file_base64 = base64.b64encode(output.read())

        self.write({"report_file": file_base64})

        return {
            "type": "ir.actions.act_url",
            "url": "/web/binary/jic_attendance_report?wizard_id=%s"
                   % (self.id),
            "target": "new",
            "tag": "reload",
        }

    def compose_periodic_attendance_mail(self, date_list, attendance_lines, employee):
        template = self.env.ref('jic_hr_attendance_report.email_template_employee_periodic_attendances', False)
        employee_id = self.env["hr.employee"].browse(employee.get("id"))
        message_composer = self.env['mail.compose.message'].with_context(
            default_use_template=bool(template),
            date_from=min(date_list),
            date_to=max(date_list),
            date_list=date_list,
            attendance_lines=attendance_lines,
            force_email=True, mail_notify_author=True
        ).create({
            'res_id': employee_id.id,
            'template_id': template and template.id or False,
            'model': 'hr.employee',
            'composition_mode': 'comment'})

        # Simulate the onchange (like trigger in form the view)
        update_values = message_composer._onchange_template_id(template.id, 'comment', 'hr.employee', employee_id.id)['value']
        message_composer.write(update_values)
        message_composer._action_send_mail()

    def _send_periodic_attendance_report(self):
        """
        Calling from cron job to send mail to employees periodically
        It must return trigger mail every 3 days with datas of last one week
        :return:
        """
        date_to = fields.Date.today()
        date_from = date_to - timedelta(days=10)

        report_wiz_id = self.create({
            "date_from": date_from,
            "date_to": date_to
        })
        date_list, attendance_lines, employees_list = report_wiz_id.get_data_to_plot()

        for employee in employees_list:
            self.compose_periodic_attendance_mail(date_list, attendance_lines, employee)


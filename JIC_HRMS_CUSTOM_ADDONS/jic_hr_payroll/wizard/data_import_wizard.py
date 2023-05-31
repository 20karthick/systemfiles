# -*- coding: utf-8 -*-
import base64
from io import BytesIO
import os
from odoo.exceptions import ValidationError

try:
    import openpyxl as xl
except ImportError:
    xl = None

from odoo import api, fields, models, _, SUPERUSER_ID


class JicDataImportWiz(models.TransientModel):
    _name = 'jic.data.import.wiz'
    _description = 'Common Data Import Wizard'

    filename = fields.Char(string="File Name")
    xlsx_file = fields.Binary(string="File", attachment=True, required=True)

    @api.model
    def check_file_requirements(self, content, filename):
        file_extension = os.path.splitext(filename)[1]
        file_size = int(len(content) * 3 / 4)  # Compute file_size in bytes
        if file_extension not in [".xlsx", ".XLSX"]:
            raise ValidationError(_("Allowed file format is xlsx"))
        if file_size > 50 * 1024 * 1024:
            raise ValidationError(_("File size must be under 5MB"))

    def create_rule(self, data):

        create_dict = data
        input_needed = False

        if data.get('category_name') == "Base":
            create_dict['category_id'] = 1
        elif data.get('category_name') == "Allowance":
            create_dict['category_id'] = 2
        else:
            create_dict['category_id'] = 4

        # Add input conditions if necessary

        if not(data.get('need_input') is None):
            create_dict.update(
                {
                    'condition_select': 'python',
                    'condition_python': 'result=inputs.' + data.get('code').strip(),
                    'amount_select': 'code',
                    'amount_python_compute': 'result=inputs.' + data.get('code').strip() + '.amount',
                }
            )
            input_needed = True

        del create_dict['category_name']
        del create_dict['need_input']
        rule_id = self.env['hr.salary.rule'].create(create_dict)

        # Prepare extra input category
        if input_needed:
            self.env['hr.employee.extra.input.category'].create(
                {
                    "name": create_dict['name'],
                    "rule_id": rule_id.id,
                }
            )
            self.env['hr.rule.input'].create(
                {
                    "name": create_dict['name'],
                    "code": data.get('code').strip(),
                    "input_id": rule_id.id,
                }
            )

    def button_create_salary_rule(self):
        self.ensure_one()
        if self.xlsx_file and xl:
            # Check file constraints
            self.check_file_requirements(self.xlsx_file, self.filename)

            # read data from excel file
            wb = xl.load_workbook(BytesIO(base64.decodebytes(self.xlsx_file)))
            ws = wb.active

            # Prepare first row
            first_row = []  # Header
            for col in range(1, ws.max_column + 1):
                first_row.append(ws.cell(1, col).value)

            # Prepare other rows based on header
            for row in range(2, ws.max_row + 1):
                elm = {}
                for col in range(1, ws.max_column + 1):
                    elm[first_row[col - 1]] = ws.cell(row, col).value

                if elm['Rule Code']:

                    self.create_rule({
                        "code": ''.join(e for e in elm['Rule Code'] if e.isalnum()),
                        "name": elm['Rule Name'],
                        "category_name": elm['Category'],
                        "amount_select": 'code',
                        "need_input": elm['Input Required'],
                    })

    # Danger ......................
    def button_update_analytic_lines(self):
        for analytic_line in self.env['account.analytic.line'].search(
            [
                ('so_line','!=',False),
                ('project_id', '!=', False),
                ('employee_id', '!=', False),
                ('amount','=',0)
            ]
        ):
            # Find cost of the employee
            employee_id = self.env['hr.employee'].browse(analytic_line.employee_id.id)
            cost_per_hour = employee_id.timesheet_cost

            amount = cost_per_hour * analytic_line.unit_amount

            analytic_line.amount = -1 * amount

    ############################################################################################
    ############################################################################################

    def create_contract_id(self, employee_id, contract_dict):
        if employee_id and contract_dict:

            journal_id = self.env['account.journal'].search([('name','=','Salary Payable')])
            if not journal_id:
                raise ValidationError(_("Salary journal is missing"))

            structure_id = self.env['hr.payroll.structure'].search([('code', '=', 'FULL TIME')])
            if not structure_id:
                raise ValidationError(_("Salary structure (FULL TIME) is missing"))

            contract_dict.update(
                {
                    'employee_id': employee_id.id,
                    'journal_id': journal_id.id,
                    'struct_id': structure_id.id
                }
            )

            return self.env['hr.contract'].create(contract_dict)

    def button_create_employee(self):
        self.ensure_one()
        if self.xlsx_file and xl:
            # Check file constraints
            self.check_file_requirements(self.xlsx_file, self.filename)

            # read data from excel file
            wb = xl.load_workbook(BytesIO(base64.decodebytes(self.xlsx_file)))
            ws = wb.active

            # Prepare first row
            first_row = []  # Header
            for col in range(1, ws.max_column + 1):
                first_row.append(ws.cell(1, col).value)

            # Prepare other rows based on header
            for row in range(2, ws.max_row + 1):
                elm = {}
                for col in range(1, ws.max_column + 1):
                    elm[first_row[col - 1]] = ws.cell(row, col).value

                if elm['Emp ID'] and elm['First Name']:

                    emp_dict = {
                        'emp_code_old': elm['Emp ID'],
                        'firstname': elm['First Name'],
                        'lastname': elm['Second Name'],
                        'lastname2': elm['Last Name'],
                        'name': elm.get('First Name') or '' + ' ' + elm.get('Second Name') or '' + ' ' + elm.get('Last Name') or '',
                        'mobile_phone': elm['Official Contact Number'],
                        'department_id': elm['First Name'],
                        'parent_id': False,
                        'work_email': elm['Official Email ID'],
                        'phone': elm['Personal Contact Number'],
                        'private_email': elm['Personal Email ID'],
                        'probation_period': 180
                    }

                    contract_dict = {
                        'name': elm['First Name'] + ' - Contract',
                        'probation_start_date': elm['Probation Start Date'],
                        'probation_end_date': elm['Probation End Date'],
                        'hr_responsible_id': False,
                        'struct_id': False,
                        'employee_id': False,
                        'journal_id': False,
                        'state': 'open',
                        'wage': 1
                    }

                    if elm['Department']:
                        department_id = self.env['hr.department'].search([('name','=', elm['Department'].strip())])
                        if department_id:
                            emp_dict['department_id'] = department_id.id
                        else:
                            department_id = self.env['hr.department'].create({'name': elm['Department'].strip()})
                            emp_dict['department_id'] = department_id.id

                    if elm['Job Title']:
                        job_id = self.env['hr.job'].search([('name', '=', elm['Job Title'].strip())])
                        if job_id:
                            emp_dict['job_id'] = job_id.id
                        else:
                            job_id = self.env['hr.job'].create({'name': elm['Job Title'].strip()})
                            emp_dict['job_id'] = job_id.id

                    print("------------------------------------------------>>>>>>>>", contract_dict)
                    employee_id = self.env['hr.employee'].create(emp_dict)

                    contract_id = self.create_contract_id(employee_id, contract_dict)

                    employee_id.contract_id = contract_id.id

                    employee_id.button_probation()

            # Loop again to fill parent
            for row in range(2, ws.max_row + 1):
                elm = {}
                for col in range(1, ws.max_column + 1):
                    elm[first_row[col - 1]] = ws.cell(row, col).value

                employee_id = self.env['hr.employee'].search([('emp_code_old','=', elm['Emp ID'])])
                parent_id = self.env['hr.employee'].search([('emp_code_old','=', elm['Reporting Manager'])])

                if elm.get('Reporting Manager') and not parent_id:
                    raise ValidationError(_("Reporting Manager code %s is missing")%(elm.get('Reporting Manager')))

                if parent_id:
                    employee_id[0].update({'parent_id': parent_id[0].id})
# -*- coding: utf-8 -*-

from datetime import datetime
from dateutil import relativedelta

from odoo import api, fields, models, _
from datetime import timedelta
from odoo.exceptions import ValidationError
from openpyxl.utils import get_column_letter
from odoo.addons.jic_base.models import openpyxl_formats as openpyxl_formats

try:
    import openpyxl as xl
except ImportError:
    xl = None
import base64
from io import BytesIO


class HrPayrollDiagnosis(models.TransientModel):
    _name = 'hr.payroll.diagnosis'
    _description = 'This is to find out all the missed before proceeding to payroll'

    report_file = fields.Binary(string="Report File")

    date_from = fields.Date(string='Date From', required=True)
    date_to = fields.Date(string='Date To', required=True)

    def plot_pending_leaves(self, wb, pending_leaves):
        sheet = wb.active
        sheet.title = "Pending Leaves"
        sheet.sheet_view.showGridLines = False
        sheet.sheet_view.windowProtection = True
        sheet.sheet_view.zoomScale = 100

        leave_state = {
            'draft': 'Draft',
            'confirm': 'To Approve',
            'validate1': 'Second Approval'
        }
        sheet.title = "Pending Leaves"

        # Set column width
        sheet.column_dimensions[get_column_letter(1)].width = 20
        sheet.column_dimensions[get_column_letter(2)].width = 20
        sheet.column_dimensions[get_column_letter(3)].width = 30
        sheet.column_dimensions[get_column_letter(4)].width = 20
        sheet.column_dimensions[get_column_letter(5)].width = 20

        sheet_row = 1
        sheet_col = 0

        sheet.cell(row=sheet_row, column=sheet_col + 1, value="Date From").style = openpyxl_formats.format_bold_no_color
        sheet.cell(row=sheet_row, column=sheet_col + 2, value="Date To").style = openpyxl_formats.format_bold_no_color
        sheet.cell(row=sheet_row, column=sheet_col + 3, value="Employee").style = openpyxl_formats.format_bold_no_color
        sheet.cell(row=sheet_row, column=sheet_col + 4, value="Leave Type").style = openpyxl_formats.format_bold_no_color
        sheet.cell(row=sheet_row, column=sheet_col + 5, value="State").style = openpyxl_formats.format_bold_no_color

        sheet_row += 1
        sheet_col = 0
        for leave in pending_leaves:
            sheet.cell(row=sheet_row, column=sheet_col + 1, value=leave.request_date_from.strftime("%d %B ,%Y")).style=openpyxl_formats.format_unbold_data_default
            sheet.cell(row=sheet_row, column=sheet_col + 2, value=leave.request_date_to.strftime("%d %B ,%Y")).style=openpyxl_formats.format_unbold_data_default
            sheet.cell(row=sheet_row, column=sheet_col + 3, value=leave.employee_id.name).style=openpyxl_formats.format_unbold_data_default
            sheet.cell(row=sheet_row, column=sheet_col + 4, value=leave.holiday_status_id.name).style=openpyxl_formats.format_unbold_data_default
            sheet.cell(row=sheet_row, column=sheet_col + 5, value=leave_state.get(leave.state)).style=openpyxl_formats.format_unbold_data_default
            sheet_row += 1
            sheet_col = 0

    def plot_pending_timesheet(self, wb, pending_timesheet):
        sheet = wb.create_sheet("Pending Time Sheets Approval")
        sheet.sheet_view.showGridLines = False
        sheet.sheet_view.windowProtection = True
        sheet.sheet_view.zoomScale = 100

        # Set column width
        sheet.column_dimensions[get_column_letter(1)].width = 20
        sheet.column_dimensions[get_column_letter(2)].width = 30
        sheet.column_dimensions[get_column_letter(3)].width = 30
        sheet.column_dimensions[get_column_letter(4)].width = 30
        sheet.column_dimensions[get_column_letter(5)].width = 20
        sheet.column_dimensions[get_column_letter(6)].width = 20

        sheet_row = 1
        sheet_col = 0

        sheet.cell(row=sheet_row, column=sheet_col + 1, value="Date").style = openpyxl_formats.format_bold_no_color
        sheet.cell(row=sheet_row, column=sheet_col + 2, value="Employee").style = openpyxl_formats.format_bold_no_color
        sheet.cell(row=sheet_row, column=sheet_col + 3, value="Project").style = openpyxl_formats.format_bold_no_color
        sheet.cell(row=sheet_row, column=sheet_col + 4, value="Task").style = openpyxl_formats.format_bold_no_color
        sheet.cell(row=sheet_row, column=sheet_col + 5, value="Hour").style = openpyxl_formats.format_bold_no_color
        sheet.cell(row=sheet_row, column=sheet_col + 6, value="Status").style = openpyxl_formats.format_bold_no_color

        sheet_row += 1
        sheet_col = 0
        for timesheet in pending_timesheet:
            sheet.cell(row=sheet_row, column=sheet_col + 1, value=timesheet.date.strftime("%d %B ,%Y")).style=openpyxl_formats.format_unbold_data_default
            sheet.cell(row=sheet_row, column=sheet_col + 2, value=timesheet.employee_id.name).style=openpyxl_formats.format_unbold_data_default
            sheet.cell(row=sheet_row, column=sheet_col + 3, value=timesheet.project_id.name).style=openpyxl_formats.format_unbold_data_default
            sheet.cell(row=sheet_row, column=sheet_col + 4, value=timesheet.task_id.name).style=openpyxl_formats.format_unbold_data_default
            sheet.cell(row=sheet_row, column=sheet_col + 5, value=timesheet.unit_amount).style=openpyxl_formats.format_unbold_data_default
            sheet.cell(row=sheet_row, column=sheet_col + 6, value=timesheet.state).style = openpyxl_formats.format_unbold_data_default
            sheet_row += 1
            sheet_col = 0

    def plot_pending_input_requests(self, wb, pending_requests):
        sheet = wb.create_sheet("Pending Payslip Input Requests")
        sheet.sheet_view.showGridLines = False
        sheet.sheet_view.windowProtection = True
        sheet.sheet_view.zoomScale = 100

        # Set column width
        sheet.column_dimensions[get_column_letter(1)].width = 20
        sheet.column_dimensions[get_column_letter(2)].width = 30
        sheet.column_dimensions[get_column_letter(3)].width = 30
        sheet.column_dimensions[get_column_letter(4)].width = 30
        sheet.column_dimensions[get_column_letter(5)].width = 20
        sheet.column_dimensions[get_column_letter(6)].width = 20
        sheet.column_dimensions[get_column_letter(7)].width = 20

        sheet_row = 1
        sheet_col = 0

        sheet.cell(row=sheet_row, column=sheet_col + 1, value="Date").style = openpyxl_formats.format_bold_no_color
        sheet.cell(row=sheet_row, column=sheet_col + 2, value="Code").style = openpyxl_formats.format_bold_no_color
        sheet.cell(row=sheet_row, column=sheet_col + 3, value="Employee").style = openpyxl_formats.format_bold_no_color
        sheet.cell(row=sheet_row, column=sheet_col + 4, value="Input").style = openpyxl_formats.format_bold_no_color
        sheet.cell(row=sheet_row, column=sheet_col + 5, value="Manager").style = openpyxl_formats.format_bold_no_color
        sheet.cell(row=sheet_row, column=sheet_col + 6, value="HR Responsible").style = openpyxl_formats.format_bold_no_color
        sheet.cell(row=sheet_row, column=sheet_col + 7, value="Amount").style = openpyxl_formats.format_bold_no_color

        sheet_row += 1
        sheet_col = 0
        for timesheet in pending_requests:
            sheet.cell(row=sheet_row, column=sheet_col + 1, value=timesheet.date.strftime("%d %B ,%Y")).style=openpyxl_formats.format_unbold_data_default
            sheet.cell(row=sheet_row, column=sheet_col + 2, value=timesheet.name).style=openpyxl_formats.format_unbold_data_default
            sheet.cell(row=sheet_row, column=sheet_col + 3, value=timesheet.employee_id.name).style=openpyxl_formats.format_unbold_data_default
            sheet.cell(row=sheet_row, column=sheet_col + 4, value=timesheet.input_category_id.name).style=openpyxl_formats.format_unbold_data_default
            sheet.cell(row=sheet_row, column=sheet_col + 5, value=timesheet.manager_id.name or " ").style=openpyxl_formats.format_unbold_data_default
            sheet.cell(row=sheet_row, column=sheet_col + 6, value=timesheet.hr_responsible_id.name or " ").style = openpyxl_formats.format_unbold_data_default
            sheet.cell(row=sheet_row, column=sheet_col + 7, value=timesheet.amount).style = openpyxl_formats.format_unbold_data_default
            sheet_row += 1
            sheet_col = 0

    def plot_pending_separation_requests(self, wb, pending_requests):
        sheet = wb.create_sheet("Pending Separation Requests")
        sheet.sheet_view.showGridLines = False
        sheet.sheet_view.windowProtection = True
        sheet.sheet_view.zoomScale = 100

        # Set column width
        sheet.column_dimensions[get_column_letter(1)].width = 20
        sheet.column_dimensions[get_column_letter(2)].width = 20
        sheet.column_dimensions[get_column_letter(3)].width = 30
        sheet.column_dimensions[get_column_letter(4)].width = 30
        sheet.column_dimensions[get_column_letter(5)].width = 20
        sheet.column_dimensions[get_column_letter(6)].width = 25
        sheet.column_dimensions[get_column_letter(7)].width = 30

        sheet_row = 1
        sheet_col = 0

        sheet.cell(row=sheet_row, column=sheet_col + 1, value="Date").style = openpyxl_formats.format_bold_no_color
        sheet.cell(row=sheet_row, column=sheet_col + 2, value="Code").style = openpyxl_formats.format_bold_no_color
        sheet.cell(row=sheet_row, column=sheet_col + 3, value="Employee").style = openpyxl_formats.format_bold_no_color
        sheet.cell(row=sheet_row, column=sheet_col + 4, value="Separation Type").style = openpyxl_formats.format_bold_no_color
        sheet.cell(row=sheet_row, column=sheet_col + 5, value="Manager").style = openpyxl_formats.format_bold_no_color
        sheet.cell(row=sheet_row, column=sheet_col + 6,
                   value="HR Responsible").style = openpyxl_formats.format_bold_no_color
        sheet.cell(row=sheet_row, column=sheet_col + 7, value="Indemnity Amount").style = openpyxl_formats.format_bold_no_color

        sheet_row += 1
        sheet_col = 0
        for timesheet in pending_requests:
            if timesheet.last_working_date_approved:
                sheet.cell(row=sheet_row, column=sheet_col + 1,
                           value=timesheet.last_working_date_approved.strftime("%d %B ,%Y")).style = openpyxl_formats.format_unbold_data_default
            else:
                sheet.cell(row=sheet_row, column=sheet_col + 1,
                           value=timesheet.date_of_request.strftime(
                               "%d %B ,%Y")).style = openpyxl_formats.format_unbold_data_default
            sheet.cell(row=sheet_row, column=sheet_col + 2,
                       value=timesheet.name).style = openpyxl_formats.format_unbold_data_default
            sheet.cell(row=sheet_row, column=sheet_col + 3,
                       value=timesheet.employee_id.name).style = openpyxl_formats.format_unbold_data_default
            sheet.cell(row=sheet_row, column=sheet_col + 4,
                       value=timesheet.checklist_type_id.name).style = openpyxl_formats.format_unbold_data_default
            sheet.cell(row=sheet_row, column=sheet_col + 5,
                       value=timesheet.manager_id.name or " ").style = openpyxl_formats.format_unbold_data_default
            sheet.cell(row=sheet_row, column=sheet_col + 6,
                       value=timesheet.hr_responsible_id.name or " ").style = openpyxl_formats.format_unbold_data_default
            sheet.cell(row=sheet_row, column=sheet_col + 7,
                       value=timesheet.settlement_amount).style = openpyxl_formats.format_unbold_data_default
            sheet_row += 1
            sheet_col = 0

    def plot_pending_overtime_requests(self, wb, pending_requests):
        sheet = wb.create_sheet("Pending Overtime Compensation Requests")
        sheet.sheet_view.showGridLines = False
        sheet.sheet_view.windowProtection = True
        sheet.sheet_view.zoomScale = 100

        # Set column width
        sheet.column_dimensions[get_column_letter(1)].width = 20
        sheet.column_dimensions[get_column_letter(2)].width = 30
        sheet.column_dimensions[get_column_letter(3)].width = 30
        sheet.column_dimensions[get_column_letter(4)].width = 30
        sheet.column_dimensions[get_column_letter(5)].width = 20
        sheet.column_dimensions[get_column_letter(6)].width = 20
        sheet.column_dimensions[get_column_letter(7)].width = 20

        sheet_row = 1
        sheet_col = 0

        sheet.cell(row=sheet_row, column=sheet_col + 1, value="Date").style = openpyxl_formats.format_bold_no_color
        sheet.cell(row=sheet_row, column=sheet_col + 2, value="Code").style = openpyxl_formats.format_bold_no_color
        sheet.cell(row=sheet_row, column=sheet_col + 3, value="Employee").style = openpyxl_formats.format_bold_no_color
        sheet.cell(row=sheet_row, column=sheet_col + 4, value="Type").style = openpyxl_formats.format_bold_no_color
        sheet.cell(row=sheet_row, column=sheet_col + 5, value="Manager").style = openpyxl_formats.format_bold_no_color
        sheet.cell(row=sheet_row, column=sheet_col + 6, value="HR Responsible").style = openpyxl_formats.format_bold_no_color
        sheet.cell(row=sheet_row, column=sheet_col + 7, value="Hours").style = openpyxl_formats.format_bold_no_color

        sheet_row += 1
        sheet_col = 0
        for timesheet in pending_requests:
            sheet.cell(row=sheet_row, column=sheet_col + 1, value=timesheet.overtime_date.strftime("%d %B ,%Y")).style=openpyxl_formats.format_unbold_data_default
            sheet.cell(row=sheet_row, column=sheet_col + 2, value=timesheet.name).style=openpyxl_formats.format_unbold_data_default
            sheet.cell(row=sheet_row, column=sheet_col + 3, value=timesheet.employee_id.name).style=openpyxl_formats.format_unbold_data_default
            sheet.cell(row=sheet_row, column=sheet_col + 4, value=timesheet.overtime_type).style=openpyxl_formats.format_unbold_data_default
            sheet.cell(row=sheet_row, column=sheet_col + 5, value=timesheet.manager_id.name or " ").style=openpyxl_formats.format_unbold_data_default
            sheet.cell(row=sheet_row, column=sheet_col + 6, value=timesheet.hr_responsible_id.name or " ").style = openpyxl_formats.format_unbold_data_default
            sheet.cell(row=sheet_row, column=sheet_col + 7, value=timesheet.amount).style = openpyxl_formats.format_unbold_data_default
            sheet_row += 1
            sheet_col = 0

    def plot_missed_attendance_and_timesheet(self, wb, date_list, attendance_lines, employees_list):
        sheet = wb.create_sheet("Missed Time Sheets and Attendance")
        sheet.sheet_view.showGridLines = False
        sheet.sheet_view.windowProtection = True
        sheet.sheet_view.zoomScale = 100

        # Set column width
        sheet.column_dimensions[get_column_letter(1)].width = 20
        sheet.column_dimensions[get_column_letter(2)].width = 30
        sheet.column_dimensions[get_column_letter(3)].width = 20
        sheet.column_dimensions[get_column_letter(4)].width = 20
        sheet.column_dimensions[get_column_letter(5)].width = 20
        sheet.column_dimensions[get_column_letter(6)].width = 20

        sheet_row = 1
        sheet_col = 0

        sheet.cell(row=sheet_row, column=sheet_col + 1, value="Date").style = openpyxl_formats.format_bold_no_color
        sheet.cell(row=sheet_row, column=sheet_col + 2, value="Employee").style = openpyxl_formats.format_bold_no_color
        sheet.cell(row=sheet_row, column=sheet_col + 3, value="Attendance Shortage").style = openpyxl_formats.format_bold_no_color
        sheet.cell(row=sheet_row, column=sheet_col + 4, value="Timesheet Required").style = openpyxl_formats.format_bold_no_color
        sheet.cell(row=sheet_row, column=sheet_col + 5, value="Timesheet Logged").style = openpyxl_formats.format_bold_no_color
        sheet.cell(row=sheet_row, column=sheet_col + 6, value="Half Day").style = openpyxl_formats.format_bold_no_color

        sheet_row += 1
        sheet_col = 0
        for line in attendance_lines:
            for e in attendance_lines[line]:
                entry = True

                emp_id = self.env['hr.employee'].browse(int(e.get('employee_id')))
                need_to_log_timesheet = emp_id.need_to_log_timesheet

                if e.get('holiday'):
                    entry = False
                if e.get('half_day'):
                    if e.get('shortage_hours') < (e.get('grace_period') / 2):
                        entry = False
                    if e.get('calendar_hours') / 2 <= e.get('approved_timesheet_hours') or not need_to_log_timesheet:
                        entry = False
                if e.get('shortage_hours') < e.get('grace_period'):
                    entry = False
                if e.get('calendar_hours') <= e.get('approved_timesheet_hours') or not need_to_log_timesheet:
                    entry = False

                if entry:

                    sheet.cell(row=sheet_row, column=sheet_col + 1,
                               value=e.get("date").strftime(
                                   "%d %B ,%Y")).style = openpyxl_formats.format_unbold_data_default
                    sheet.cell(row=sheet_row, column=sheet_col + 2,
                               value=e.get("employee_name")).style = openpyxl_formats.format_unbold_data_default
                    sheet.cell(row=sheet_row, column=sheet_col + 3,
                               value=e.get("shortage_hours_str")).style = openpyxl_formats.format_unbold_data_default
                    sheet.cell(row=sheet_row, column=sheet_col + 4,
                               value='4:0' if e.get('half_day') else e.get("calendar_hours_str")).style = openpyxl_formats.format_unbold_data_default
                    sheet.cell(row=sheet_row, column=sheet_col + 5,
                               value=e.get("approved_timesheet")).style = openpyxl_formats.format_unbold_data_default
                    sheet.cell(row=sheet_row, column=sheet_col + 6,
                               value='True' if e.get("half_day") else '').style = openpyxl_formats.format_unbold_data_default

                    sheet_row += 1
                    sheet_col = 0

    def do_diagnosis(self):
        '''
        1. Open Leave Requests
        2. Un approved Time sheets
        3. Attendance mismatch
        4. Timesheet mismatch
        :return:
        '''
        employee_ids = False
        date_from = self.date_from
        date_to = self.date_to
        days = [date_from+timedelta(days=1) for x in range((date_to-date_from).days)]
        wb = xl.Workbook()

        employee_ids = self.env['hr.employee'].search([])

        ###############################
        # Open Leave Requests
        ##############################

        domain = ['|','&',
            ('employee_id','in',employee_ids.ids),
            '&',
            ('request_date_from','>=', date_from),
            ('request_date_from','<=', date_to),
            '&',
            ('employee_id', 'in', employee_ids.ids),
            '&',
            ('request_date_to', '>=', date_from),
            ('request_date_to', '<=', date_to)
        ]

        pending_leaves = self.env['hr.leave'].search(domain).filtered(
            lambda a: a.state in ['draft', 'confirm', 'validate1']
        )
        self.plot_pending_leaves(wb, pending_leaves)

        ###############################
        # Pending Timesheet
        ##############################

        domain = [
            ('date','>=', date_from),
            ('date','<=', date_to),
            ('state','in', ['submitted'])
        ]
        pending_timesheets = self.env['hr.timesheet.entry'].search(domain)
        self.plot_pending_timesheet(wb, pending_timesheets)

        ###############################
        # Missed Leaves and Time sheets
        ##############################
        report_wiz_id = self.env['hr.attendance.report.wizard'].create({
            "date_from": date_from,
            "date_to": date_to,
        })
        date_list, attendance_lines, employees_list = report_wiz_id.get_data_to_plot()
        self.plot_missed_attendance_and_timesheet(wb, date_list, attendance_lines, employees_list)

        ###############################
        # Pending Payslip Input Requests
        ##############################
        domain = [
            ('date', '>=', date_from),
            ('date', '<=', date_to),
            ('state', 'in', ['requested', 'validated'])
        ]
        pending_requests = self.env['hr.employee.input.requests'].search(domain)
        self.plot_pending_input_requests(wb, pending_requests)

        ###############################
        # Pending Separation Requests
        ##############################
        domain = [
            ('last_working_date_approved', '>=', date_from),
            ('last_working_date_approved', '<=', date_to),
            ('state', 'not in', ['completed', 'draft'])
        ]
        pending_requests = self.env['hr.employee.separation'].search(domain)
        self.plot_pending_separation_requests(wb, pending_requests)

        ##############################
        # Pending Overtime Requests
        ##############################
        domain = [
            ('overtime_date', '>=', date_from),
            ('overtime_date', '<=', date_to),
            ('state', 'in', ['requested', 'validated'])
        ]
        pending_overtime = self.env['hr.overtime.request'].search(domain)
        self.plot_pending_overtime_requests(wb, pending_overtime)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        file_base64 = base64.b64encode(output.read())
        self.write({"report_file": file_base64})

        return {
            "type": "ir.actions.act_url",
            "url": "/web/binary/jic_payroll_diagnosis?wizard_id=%s"
                   % (self.id),
            "target": "new",
            "tag": "reload",
        }




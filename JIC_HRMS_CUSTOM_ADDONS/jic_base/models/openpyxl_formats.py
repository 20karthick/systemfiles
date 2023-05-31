from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side

format_unbold_data_default = NamedStyle(name="format_unbold_data_default")
format_unbold_data_default.font = Font(bold=False, size=10, color="000000")
format_unbold_data_default.alignment = Alignment(horizontal="center", vertical="center")
format_unbold_data_default.number_format = "0.00"
format_unbold_data_default.border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

format_leave = NamedStyle(name="format_leave")
format_leave.font = Font(bold=False, size=10, color="000000")
format_leave.alignment = Alignment(horizontal="center", vertical="center")
format_leave.number_format = "0.00"
format_leave.fill = PatternFill("solid", start_color="ff0000")
format_leave.border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

format_bold_no_color = NamedStyle(name="format_bold_no_color")
format_bold_no_color.font = Font(bold=True, size=10, color="000000")
format_bold_no_color.alignment = Alignment(horizontal="center", vertical="center")
format_bold_no_color.number_format = "0.00"
format_bold_no_color.border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

format_bold_no_color_special = NamedStyle(name="format_bold_no_color_special")
format_bold_no_color_special.font = Font(bold=True, size=10, color="000000")
format_bold_no_color_special.alignment = Alignment(horizontal="center", vertical="center")
format_bold_no_color_special.number_format = "0.00"


format_bold_with_color = NamedStyle(name="format_bold_with_color")
format_bold_with_color.font = Font(bold=True, size=10, color="000000")
format_bold_with_color.alignment = Alignment(horizontal="center", vertical="center")
format_bold_with_color.number_format = "0.00"
format_bold_with_color.fill = PatternFill("solid", start_color="5cb800")

format_unbold_with_color = NamedStyle(name="format_unbold_with_color")
format_unbold_with_color.font = Font(bold=False, size=10, color="000000")
format_unbold_with_color.alignment = Alignment(horizontal="center", vertical="center")
format_unbold_with_color.number_format = "0.00"
format_unbold_with_color.fill = PatternFill("solid", start_color="b4c7dc")

format_unbold_with_color_danger = NamedStyle(name="format_unbold_with_color_danger")
format_unbold_with_color_danger.font = Font(bold=True, size=10, color="0000ff")
format_unbold_with_color_danger.alignment = Alignment(horizontal="center", vertical="center")
format_unbold_with_color_danger.number_format = "0.00"
#format_unbold_with_color_danger.fill = PatternFill("solid", start_color="3465a4")
format_unbold_with_color_danger.border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

format_vaccant_row = NamedStyle(name="format_vaccant_row")
format_vaccant_row.fill = PatternFill("solid", start_color="b4c7dc")
format_vaccant_row.border = Border(
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

format_table_head = NamedStyle(name="format_table_head")
format_table_head.font = Font(bold=True, size=10, color="000000")
format_table_head.alignment = Alignment(horizontal="center", vertical="center")
format_table_head.fill = PatternFill("solid", start_color="5cb800")

format_final_counts = NamedStyle(name="format_final_counts")
format_final_counts.font = Font(bold=True, size=11, color="000000")
format_final_counts.alignment = Alignment(horizontal="center", vertical="center")


# From reference

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
format_date = NamedStyle(name="format_date", number_format="DD-MM-YYYY")
format_date.font = Font(bold=False, size=10, color="000000")
format_date.border = thin_border
format_date.alignment = Alignment(horizontal="left", vertical="center")

# Title
format_title = NamedStyle(name="format_title")
format_title.fill = PatternFill(
    start_color="FFF58C", end_color="FFF58C", fill_type="solid"
)
format_title.font = Font(bold=True, size=11, color="000000")
format_title.alignment = Alignment(horizontal="center", vertical="center")

# Headers Style
format_header = NamedStyle(name="format_header")
format_header.font = Font(bold=True, size=10, color="000000")
format_header.border = thin_border

# Filter Value
format_filter_value = NamedStyle(name="format_filter_value")
format_filter_value.font = Font(bold=False, size=10, color="000000")
format_filter_value.border = thin_border

# Content Table Header
format_content_header = NamedStyle(name="format_content_header")
format_content_header.fill = PatternFill(
    start_color="ccdefc", end_color="ccdefc", fill_type="solid"
)
format_content_header.font = Font(bold=True, size=10, color="000000")
format_content_header.border = thin_border

# Content Line
format_line = NamedStyle(name="format_line")
format_line.font = Font(bold=False, size=10, color="000000")
format_line.border = thin_border